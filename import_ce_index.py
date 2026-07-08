"""
Import county code enforcement index files into property_records.db.
Aggregates all unique CaseNo -> APN -> Address -> City mappings.
"""

import openpyxl, os, sqlite3
from pathlib import Path

SCRIPT_DIR = Path(__file__).resolve().parent
STATE_DIR = SCRIPT_DIR / "state"
PROPERTY_DB = STATE_DIR / "property_records.db"
COURT_DB = STATE_DIR / "court_calendar.db"

INDEX_DIRS = [
    r"C:\Users\mark\Documents\code enforcement index files",
    r"C:\Users\mark\Documents\code enforcement writ",
    r"C:\Users\mark\Documents\code enforcement case list and files",
]


def get_all_xlsx():
    seen = set()
    for base in INDEX_DIRS:
        if not os.path.isdir(base):
            continue
        for f in os.listdir(base):
            if f.endswith(".xlsx") and f not in seen:
                seen.add(f)
                yield os.path.join(base, f)


def get_field_map(headers):
    return {
        "caseno": next((i for i, h in enumerate(headers) if h == "CaseNo"), 0),
        "apn": next((i for i, h in enumerate(headers) if h == "APN"), None),
        "address": next((i for i, h in enumerate(headers) if h == "Site Address"), None),
        "city": next((i for i, h in enumerate(headers) if "city" in h.lower() or "community" in h.lower()), None),
        "status": next((i for i, h in enumerate(headers) if "status" in h.lower()), None),
    }


def main():
    print("Importing code enforcement index files...")
    pconn = sqlite3.connect(str(PROPERTY_DB))
    pconn.execute("PRAGMA journal_mode=WAL")
    cconn = sqlite3.connect(str(COURT_DB))

    seen_apns = set()
    seen_cases = set()
    imported = 0
    case_links = []

    for fpath in get_all_xlsx():
        fname = os.path.basename(fpath)
        wb = openpyxl.load_workbook(fpath, read_only=True, data_only=True)
        ws = wb.active
        headers = [str(c.value) for c in next(ws.iter_rows(min_row=1, max_row=1))]
        fm = get_field_map(headers)
        if fm["apn"] is None:
            wb.close()
            continue

        for row in ws.iter_rows(min_row=2, values_only=True):
            if len(row) <= max(fm["apn"], fm["address"] or 0, fm["city"] or 0):
                continue

            apn = str(row[fm["apn"]] or "").strip()
            addr = str(row[fm["address"]] or "").strip() if fm["address"] is not None else ""
            city = str(row[fm["city"]] or "").strip() if fm["city"] is not None else ""
            caseno = str(row[fm["caseno"]] or "").strip()
            status = str(row[fm["status"]] or "").strip() if fm["status"] is not None else ""

            if not apn and not addr:
                continue

            apn_norm = apn.replace("-", "").replace(" ", "")
            if apn_norm and apn_norm not in seen_apns:
                seen_apns.add(apn_norm)
                pconn.execute(
                    "INSERT INTO property_addresses (apn, street_number, street_name, city, source) VALUES (?, ?, ?, ?, ?)",
                    (apn_norm, "", addr, city, f"code_enforcement_index:{fname}"),
                )
                imported += 1
            elif not apn_norm and addr:
                pconn.execute(
                    "INSERT INTO property_addresses (street_number, street_name, city, source) VALUES (?, ?, ?, ?)",
                    ("", addr, city, f"code_enforcement_index:{fname}"),
                )
                imported += 1

            if caseno and apn_norm:
                link_key = (caseno, apn_norm)
                if link_key not in seen_cases:
                    seen_cases.add(link_key)
                    case_links.append((apn_norm, caseno, f"code_enforcement:{fname}", f"From {fname} ({status})"))

        wb.close()

    pconn.commit()

    # Insert case_property_links into court_calendar.db
    cconn.execute("PRAGMA journal_mode=WAL")
    for apn, caseno, link_type, notes in case_links:
        cconn.execute(
            "INSERT INTO case_property_links (apn, case_number, link_type, notes) VALUES (?, ?, ?, ?)",
            (apn, caseno, link_type, notes),
        )
    cconn.commit()

    print(f"  Imported {imported} property records")
    print(f"  Created {len(case_links)} case_property_links")
    print(f"  Unique APNs: {len(seen_apns)}")

    pconn.close()
    cconn.close()


if __name__ == "__main__":
    main()
